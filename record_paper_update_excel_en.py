import argparse
from datetime import datetime
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description='Append one update row into the Excel UpdateLog sheet')
parser.add_argument('--file', default='paper_reading_summary_en.xlsx', help='Excel file path')
parser.add_argument('--update', required=True, help='Update item')
parser.add_argument('--note', default='Content updated', help='Notes for this update')
args = parser.parse_args()

wb = load_workbook(args.file)
ws = wb['UpdateLog'] if 'UpdateLog' in wb.sheetnames else wb.create_sheet('UpdateLog')

if ws.max_row == 1 and ws['A1'].value is None:
    ws.append(['Timestamp (Asia/Tokyo)','Update Item','Notes'])

if ws.max_row == 0:
    ws.append(['Timestamp (Asia/Tokyo)','Update Item','Notes'])

if ws['A1'].value != 'Timestamp (Asia/Tokyo)':
    ws.insert_rows(1)
    ws['A1'] = 'Timestamp (Asia/Tokyo)'
    ws['B1'] = 'Update Item'
    ws['C1'] = 'Notes'

now = datetime.now().astimezone().strftime('%Y-%m-%d %H:%M:%S %z')
if len(now) > 5:
    now = now[:-2] + ':' + now[-2:]

ws.append([now, args.update, args.note])
wb.save(args.file)
print(f'Logged update at {now}')
